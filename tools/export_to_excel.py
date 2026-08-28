"""Jadvalni Excel faylga export qilish — avval generatsiya, keyin export"""
import sys, os, time
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.db_manager import DatabaseManager
from core.scheduler import TimetableScheduler

db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "smartdj_test.db")
db = DatabaseManager()
db.db_name = db_path
db.initialize()

classes = db.get_all_classes()
print(f"📊 {len(classes)} sinf")

# 1. Jadval generatsiya qilish
print("\n🚀 Jadval generatsiya qilinmoqda...")
scheduler = TimetableScheduler(algorithm="brkga")
start = time.time()
all_data, conflicts = scheduler.generate_all_class_timetables(
    classes, db, cancel_flag=lambda: False,
    progress_callback=lambda cn, i, t, s: None if i % 50 else print(f"   {i}/{t}")
)
elapsed = time.time() - start
print(f"   ⏱️ {elapsed:.1f}s, {len(all_data)} dars, {len(conflicts)} ziddiyat")

# 2. Bazaga saqlash
print("\n💾 Bazaga saqlanmoqda...")
db.save_scheduled_lessons(all_data, week_index=0)

# 3. Excel ga export
print("\n📊 Excel ga export qilinmoqda...")
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

KUNLAR = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba"]

wb = Workbook()
wb.remove(wb.active)

# Umumiy jadval (barcha sinflar)
ws_all = wb.create_sheet("Umumiy jadval")
ws_all.cell(1, 1, "UMUMIY DARS JADVALI — 250 sinf").font = Font(bold=True, size=14)
ws_all.cell(3, 1, "Sinf").font = Font(bold=True, color="FFFFFF")
ws_all.cell(3, 1).fill = PatternFill(start_color="2C3E50", fill_type="solid")

col = 2
for di, kun in enumerate(KUNLAR):
    ws_all.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 5)
    c = ws_all.cell(3, col, kun)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="2C3E50", fill_type="solid")
    c.alignment = Alignment(horizontal='center')
    for pp in range(6):
        cc = ws_all.cell(4, col + pp, str(pp + 1))
        cc.font = Font(bold=True, color="FFFFFF", size=8)
        cc.fill = PatternFill(start_color="34495E", fill_type="solid")
        cc.alignment = Alignment(horizontal='center')
    col += 6

for ci, cls in enumerate(classes):
    rn = ci + 5
    ws_all.cell(rn, 1, cls[1]).font = Font(bold=True, color="FFFFFF", size=8)
    ws_all.cell(rn, 1).fill = PatternFill(start_color="34495E", fill_type="solid")
    ws_all.cell(rn, 1).alignment = Alignment(horizontal='center')
    for day in range(6):
        for period in range(6):
            info = all_data.get((cls[0], day, period), {})
            val = info.get('subject_name', '') if info else ''
            c = ws_all.cell(rn, day * 7 + period + 2, val)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.font = Font(size=7)

ws_all.column_dimensions['A'].width = 8
for col in range(2, 44):
    ws_all.column_dimensions[get_column_letter(col)].width = 10

# Saqlash
output = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "jadval_250.xlsx")
wb.save(output)
print(f"\n✅ Excel saqlandi: {output}")

db.close()
