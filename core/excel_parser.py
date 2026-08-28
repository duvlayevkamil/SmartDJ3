"""
Tayanch Reja Excel Parser — Excel fayldan tayanch reja jadvallarini parse qilish.
"""

import re
from openpyxl import load_workbook


class TayanchRejaExcelParser:
    """Excel fayldan tayanch reja jadvalini parse qilish."""

    def __init__(self):
        self.errors = []

    def parse(self, excel_path):
        """
        Excel fayldan barcha tayanch reja jadvallarini topadi.
        Returns: dict — {
            "subjects": [{"name": str, "short": str, "is_group": bool}],
            "classes": [{"name": str, "level": int}],
            "hours": [[int, ...], ...]
        }
        """
        self.errors = []

        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            if ws is None:
                self.errors.append("Excel faylda jadval topilmadi!")
                return None

            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append(list(row))

            if not all_rows:
                self.errors.append("Excel faylda ma'lumot topilmadi!")
                return None

            return self._parse_rows(all_rows)

        except Exception as e:
            self.errors.append(f"Excel o'qishda xatolik: {str(e)}")
            return None

    def _parse_rows(self, rows):
        """Barcha qatorlarni parse qilish."""
        subjects = []
        hours = []
        classes = []
        header_found = False

        for row in rows:
            if not row or len(row) < 3:
                continue

            # Sarlavha qatorlarini tashlab o'tish
            if self._is_header_row(row):
                if not header_found:
                    classes = self._extract_class_columns(row)
                    if classes:
                        header_found = True
                continue

            # Agar header hali topilmasa — sinf raqamlarini qidirish
            if not header_found:
                classes = self._extract_class_columns(row)
                if classes:
                    header_found = True
                continue

            # "Jami soat" qatorini tashlab o'tish
            first_cell = self._clean(str(row[0])) if row[0] else ""
            second_cell = self._clean(str(row[1])) if row[1] else ""
            combined = first_cell + " " + second_cell
            if 'jami' in combined.lower():
                continue

            # Fan nomini topish
            subject_name = second_cell if second_cell else first_cell
            if not subject_name or len(subject_name) < 2:
                continue
            if re.match(r'^\d+$', subject_name):
                continue

            skip_words = ['nomlari', 'haftalik', 'umumiy', 'soat', 'sinflar',
                          't/r', 'fan yo', 'va nomlari']
            if any(sw in subject_name.lower() for sw in skip_words):
                continue

            # Guruh sarlavhasimi?
            is_group = self._is_group_header(subject_name)

            # Soatlar — ustun 2 dan boshlab
            row_hours = []
            for col_idx in range(2, len(row)):
                row_hours.append(self._parse_hours(row[col_idx]))

            # Agar yetarli ustun bo'lmasa — qoldiqni 0 bilan to'ldirish
            while len(row_hours) < 11:
                row_hours.append(0)

            short = "" if is_group else self._make_short(subject_name)
            subjects.append({"name": subject_name, "short": short, "is_group": is_group})
            hours.append(row_hours[:11])  # Faqat 11 sinf

        if not classes:
            classes = [{"name": f"{i}-sinf", "level": i} for i in range(1, 12)]

        if not subjects:
            self.errors.append("Fanlar topilmadi!")
            return None

        return {
            "subjects": subjects,
            "classes": classes,
            "hours": hours,
        }

    def _is_header_row(self, row):
        """Sarlavha qatormi?"""
        text = " ".join(self._clean(str(c)) for c in row if c)
        text_lower = text.lower()
        keywords = ['fan', 'soat', 't/r', 'nomlari', 'sinflar', 'haftalik', 'umumiy']
        return sum(1 for kw in keywords if kw in text_lower) >= 2

    def _extract_class_columns(self, row):
        """Sarlavha qatoridan sinf ustunlarini aniqlash."""
        classes = []
        for i, cell in enumerate(row):
            text = self._clean(str(cell)) if cell else ""
            m = re.match(r'^(\d{1,2})$', text)
            if m:
                level = int(m.group(1))
                if 1 <= level <= 11:
                    classes.append({"name": f"{level}-sinf", "level": level})
        return classes

    def _is_group_header(self, text):
        """Guruh sarlavhasi?"""
        t = text.strip().lower()
        if re.match(r'^[ivx]+\.?\s', t):
            return True
        if re.search(r'\bfanlar[si]?$', t):
            return True
        return False

    def _clean(self, text):
        """Matnni tozalash."""
        if not text:
            return ""
        text = str(text).strip()
        text = re.sub(r'\s+', ' ', text)
        text = text.replace('\n', ' ').replace('\r', '')
        return text.strip()

    def _parse_hours(self, cell):
        """Katakdan soat."""
        if cell is None:
            return 0
        text = self._clean(str(cell))
        if not text or text in ('-', '—'):
            return 0
        text = text.replace(',', '.')
        m = re.search(r'(\d+\.?\d*)', text)
        if m:
            val = float(m.group(1))
            return int(val) if val == int(val) else val
        return 0

    def _make_short(self, name):
        """Qisqartma nom."""
        words = name.split()
        if len(words) == 1:
            return name[:4].title()
        return ''.join(w[0].upper() for w in words[:3])

    def parse_for_display(self, excel_path):
        """Parse qilib, flat formatga o'tkazish."""
        result = self.parse(excel_path)
        if not result:
            return None

        flat = []
        for i, subj in enumerate(result['subjects']):
            for j, cls in enumerate(result['classes']):
                h = result['hours'][i][j] if i < len(result['hours']) and j < len(result['hours'][i]) else 0
                flat.append({
                    'subject_name': subj['name'],
                    'subject_short': subj.get('short', ''),
                    'class_level': cls['level'],
                    'weekly_hours': h,
                    'is_group': subj.get('is_group', False),
                })
        return flat
